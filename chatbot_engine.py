#!/usr/bin/env python3
"""
SwasthyaMitra — Chatbot Engine v5 (Local RAG)
• 100% free & offline — no paid APIs, no external HTTP calls.
• BM25 ranking with spelling correction, synonym expansion, and
  phrase-level matching.
• Reads PDFs (pypdf) and Excel (openpyxl) from the Doctor's Data folder.
• Provides HTML-formatted responses for the frontend chatbot widget.
"""

import argparse
import json
import math
import os
import re
import random
from collections import Counter, defaultdict
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple

from openpyxl import load_workbook
from pypdf import PdfReader


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

DATA_DIR = Path(__file__).parent / "Doctor's Data"
INDEX_FILE = Path(__file__).parent / "index_store.json"
MAX_PREVIEW_CHARS = 900
INDEX_SCHEMA_VERSION = 2

INTENT_WORDS = {
    "address", "location", "where", "near", "nearby", "around",
    "phone", "mobile", "contact", "number", "email", "details",
}

ORG_HINT_WORDS = {
    "dental", "clinic", "care", "center", "centre", "hospital",
    "medical", "health", "lab", "laboratory",
}

STOPWORDS = {
    "a", "an", "and", "are", "as", "at", "be", "by", "for", "from",
    "has", "he", "in", "is", "it", "its", "of", "on", "that", "the",
    "to", "was", "were", "will", "with", "can", "you", "me", "my",
    "i", "we", "our", "or", "find", "show", "tell", "list", "same",
    "doctor", "doctors", "hospital", "hospitals",
}

# Synonym expansion for medical specialties
_SYNONYMS: Dict[str, List[str]] = {
    "eye":        ["ophthalmologist", "opthal", "cataract", "retina", "glaucoma"],
    "eyes":       ["ophthalmologist", "opthal", "cataract"],
    "skin":       ["dermatologist", "dermatology"],
    "heart":      ["cardiologist", "cardiac", "cardiology", "cardiovascular"],
    "bone":       ["orthopedic", "ortho", "orthopaedic", "joint", "spine"],
    "joints":     ["orthopedic", "ortho", "joint", "replacement"],
    "child":      ["pediatrician", "pediatric", "paediatric"],
    "children":   ["pediatrician", "pediatric", "paediatric"],
    "kids":       ["pediatrician", "pediatric"],
    "teeth":      ["dentist", "dental", "orthodontist"],
    "tooth":      ["dentist", "dental"],
    "brain":      ["neurologist", "neurology", "neuro"],
    "nerve":      ["neurologist", "neurology"],
    "kidney":     ["nephrologist", "nephrology", "urology", "urologist", "dialysis"],
    "ear":        ["ent", "audiologist"],
    "nose":       ["ent"],
    "throat":     ["ent"],
    "ent":        ["ear", "nose", "throat"],
    "women":      ["gynecologist", "gynaecologist", "obstetrician", "gynecology"],
    "ladies":     ["gynecologist", "gynaecologist"],
    "pregnancy":  ["gynecologist", "gynaecologist", "obstetrician", "maternity"],
    "sugar":      ["diabetologist", "endocrinologist", "diabetes"],
    "diabetes":   ["diabetologist", "endocrinologist"],
    "cancer":     ["oncologist", "oncology"],
    "stomach":    ["gastroenterologist", "gastro", "gastrology"],
    "digestion":  ["gastroenterologist", "gastro"],
    "liver":      ["gastroenterologist", "hepatologist"],
    "lung":       ["pulmonologist", "chest", "respiratory"],
    "breathing":  ["pulmonologist", "chest", "respiratory"],
    "mental":     ["psychiatrist", "psychologist", "psychiatry"],
    "depression": ["psychiatrist", "psychologist"],
    "blood":      ["hematologist", "pathologist", "pathology"],
    "fever":      ["physician", "general"],
    "general":    ["physician", "gp"],
    "xray":       ["radiologist", "radiology", "imaging"],
    "scan":       ["radiologist", "radiology", "ct", "mri"],
    "plastic":    ["cosmetic"],
    "cosmetic":   ["plastic"],
    "urine":      ["urologist", "urology", "nephrologist"],
    "thyroid":    ["endocrinologist", "endocrinology"],
    "bp":         ["cardiologist", "physician"],
}


# ---------------------------------------------------------------------------
# Text helpers
# ---------------------------------------------------------------------------

def normalize_text(text: str) -> str:
    text = text.lower().strip()
    text = re.sub(r"\s+", " ", text)
    return text


def tokenize(text: str) -> List[str]:
    text = normalize_text(text)
    tokens = re.findall(r"[a-z0-9]+", text)
    return [t for t in tokens if t not in STOPWORDS and len(t) > 1]


def _expand_synonyms(tokens: List[str]) -> List[str]:
    """Add synonym tokens for medical specialty queries."""
    seen = set(tokens)
    extra = []
    for t in tokens:
        for syn in _SYNONYMS.get(t, []):
            if syn not in seen:
                extra.append(syn)
                seen.add(syn)
    return tokens + extra


def is_greeting(text: str) -> bool:
    cleaned = normalize_text(text)
    return cleaned in {"hi", "hello", "hey", "hii", "yo"}


def _esc(text: str) -> str:
    """Escape HTML special characters."""
    return text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def _phone_span(text: str) -> str:
    """Highlight phone/mobile numbers in green."""
    return re.sub(
        r"(\b\d[\d\s\(\)\-\.]{6,}\d\b)",
        r'<span style="color:#2e7d32;font-weight:500;">📞 \1</span>',
        text,
    )


# ---------------------------------------------------------------------------
# Small-talk
# ---------------------------------------------------------------------------

def get_smalltalk_reply(text: str) -> str:
    cleaned = normalize_text(text)
    words = set(re.findall(r"[a-z0-9]+", cleaned))

    greeting_tokens = {"hi", "hii", "hello", "hey", "yo"}
    morning_tokens = {"good", "morning"}
    afternoon_tokens = {"good", "afternoon"}
    evening_tokens = {"good", "evening"}
    night_tokens = {"good", "night"}
    thanks_tokens = {"thanks", "thank", "thankyou"}
    bye_tokens = {"bye", "by", "goodbye", "see", "later"}
    ok_tokens = {"ok", "okay", "k"}

    if words & greeting_tokens:
        return (
            "Hello! 👋 I'm <b>SwasthyaBot</b>, your medical assistant.<br><br>"
            "I can help you find:<br>"
            "🏥 Hospitals in Gujarat<br>"
            "👨‍⚕️ Doctors in Ahmedabad, Mumbai & other Gujarat cities<br>"
            "🏨 AMC Urban Health Centers<br><br>"
            "Try asking:<br>"
            "<i>\"Eye specialist in Ahmedabad\"</i><br>"
            "<i>\"Orthopedic hospital in Surat\"</i><br>"
            "<i>\"Health centers near Maninagar\"</i>"
        )
    if morning_tokens.issubset(words):
        return "Good morning! ☀️ Ask me anything about doctors, hospitals, locations, phones, or addresses."
    if afternoon_tokens.issubset(words):
        return "Good afternoon! 🌤️ I can help you find doctors and hospitals from your uploaded data."
    if evening_tokens.issubset(words):
        return "Good evening! 🌙 Share your query and I will fetch the best match."
    if night_tokens.issubset(words):
        return "Good night! 🌙 I'm here whenever you want to continue searching your data."
    if "doing" in words and ("what" in words or "u" in words or "you" in words):
        return "I'm here and working! 💪 I can search your uploaded data and return grounded answers."
    if "how" in words and "you" in words and ("are" in words or "r" in words):
        return "I'm doing well and ready to help! 😊 Ask me any doctor or hospital query."
    if "who" in words and "you" in words:
        return "I'm <b>SwasthyaBot</b>, your local data chatbot. I search your PDF and Excel files to answer queries. 🤖"
    if ("can" in words and "do" in words) or ("help" in words and ("what" in words or "how" in words)):
        return (
            "I can search both PDF and Excel data for doctors/hospitals by <b>name, area, city, "
            "phone, address, specialty</b>, and related natural-language queries."
        )
    if words & thanks_tokens:
        return "You're welcome! 😊 Ask another query anytime."
    if words & ok_tokens and len(words) <= 3:
        return "Okay! Share your next query whenever you're ready. 👍"
    if "what" in words and "can" in words and "you" in words and "do" in words:
        return (
            "I can fetch information from your files and answer in natural language. "
            "Try queries like: <i>doctor near Andheri</i>, <i>address of Sterling Hospital</i>, <i>phone of Dr X</i>."
        )
    if words & bye_tokens:
        return "Goodbye! 👋 Stay healthy! 🌿"
    return ""


# ---------------------------------------------------------------------------
# Chunk dataclass
# ---------------------------------------------------------------------------

@dataclass
class Chunk:
    chunk_id: int
    source_file: str
    source_type: str
    location: str
    text: str


# ---------------------------------------------------------------------------
# Local RAG Chatbot (core engine)
# ---------------------------------------------------------------------------

class LocalRAGChatbot:
    def __init__(self, data_dir: Path, index_file: Path):
        self.data_dir = data_dir
        self.index_file = index_file
        self.chunks: List[Chunk] = []
        self.doc_freq: Dict[str, int] = {}
        self.avg_doc_len: float = 0.0
        self.tokenized_chunks: List[List[str]] = []
        self.chunk_term_freqs: List[Dict[str, int]] = []
        self._all_known_terms: set = set()

    def build_index(self) -> None:
        self.chunks = []
        chunk_id = 0

        if not self.data_dir.exists():
            raise FileNotFoundError(f"Data directory not found: {self.data_dir}")

        for file_path in sorted(self.data_dir.iterdir()):
            suffix = file_path.suffix.lower()
            if suffix == ".pdf":
                pdf_chunks = self._extract_pdf_chunks(file_path, start_id=chunk_id)
                self.chunks.extend(pdf_chunks)
                chunk_id += len(pdf_chunks)
            elif suffix in {".xlsx", ".xlsm"}:
                xlsx_chunks = self._extract_xlsx_chunks(file_path, start_id=chunk_id)
                self.chunks.extend(xlsx_chunks)
                chunk_id += len(xlsx_chunks)

        if not self.chunks:
            raise ValueError("No extractable chunks found from PDF/XLSX files in Data folder.")

        self._prepare_bm25_structures()
        self._save_index()

    def load_or_build_index(self, force_rebuild: bool = False) -> None:
        if force_rebuild or not self.index_file.exists() or self._needs_rebuild():
            self.build_index()
            return
        try:
            self._load_index()
        except Exception:
            self.build_index()

    def _needs_rebuild(self) -> bool:
        if not self.index_file.exists():
            return True
        if not self._index_version_matches():
            return True
        try:
            index_mtime = self.index_file.stat().st_mtime
        except OSError:
            return True

        if not self.data_dir.exists():
            return False

        for file_path in self.data_dir.iterdir():
            if file_path.suffix.lower() not in {".pdf", ".xlsx", ".xlsm"}:
                continue
            try:
                if file_path.stat().st_mtime > index_mtime:
                    return True
            except OSError:
                continue
        return False

    def _index_version_matches(self) -> bool:
        try:
            payload = json.loads(self.index_file.read_text(encoding="utf-8"))
            return payload.get("schema_version") == INDEX_SCHEMA_VERSION
        except Exception:
            return False

    def ask(self, query: str, top_k: int = 5) -> str:
        """Return an HTML-formatted answer for the frontend chatbot."""
        if not self.chunks:
            raise ValueError("Index is not ready. Build or load index first.")

        smalltalk = get_smalltalk_reply(query)
        if smalltalk:
            return smalltalk

        if is_greeting(query):
            return "Hello! 👋 Ask me about doctors or hospitals by name, city, specialty, or area."

        query_terms = tokenize(query)
        if not query_terms:
            return "Please ask with a few specific keywords, e.g. doctor name, area, or specialty. 🔍"

        wants_address = self._wants_address(query_terms)
        retrieval_terms = [t for t in query_terms if t not in INTENT_WORDS]
        if not retrieval_terms:
            retrieval_terms = query_terms

        # Expand synonyms for medical specialty matching
        retrieval_terms = _expand_synonyms(retrieval_terms)

        corrected_query_terms, corrected_pairs = self._correct_terms(retrieval_terms)
        required_terms = self._required_terms(corrected_query_terms)
        phrase_filters = self._phrase_filters(query)
        scored = self._retrieve(
            corrected_query_terms,
            required_terms=required_terms,
            phrase_filters=phrase_filters,
            top_k=top_k,
        )

        if not scored:
            return (
                '🤔 No results found for <b>' + _esc(query) + '</b>.<br><br>'
                '💡 Try asking like:<br>'
                '• <i>"Eye specialist in Ahmedabad"</i><br>'
                '• <i>"Orthopedic hospitals in Surat"</i><br>'
                '• <i>"AMC health centers near Maninagar"</i><br>'
                '• <i>"Cardiologist in Rajkot"</i>'
            )

        # Build HTML response
        html_parts = []

        if corrected_pairs:
            fixed = ", ".join([f"{old}→{new}" for old, new in corrected_pairs])
            html_parts.append(
                f'<small style="color:#888;">🔧 Spelling normalization: {_esc(fixed)}</small><br>'
            )

        html_parts.append(
            f'Here are <b>{len(scored)}</b> result(s) for '
            f'<b>{_esc(query)}</b> '
            f'<small style="color:#888;">(🔍 Medical Data)</small>:<br><br>'
        )

        for rank, (score, coverage, chunk) in enumerate(scored, start=1):
            preview = re.sub(r"\s+", " ", chunk.text).strip()

            if wants_address:
                address = self._extract_address_like_text(preview)
                if address:
                    preview = f"Address: {address}"

            if len(preview) > MAX_PREVIEW_CHARS:
                preview = preview[:MAX_PREVIEW_CHARS] + "..."

            # Format each result as an HTML card
            source_label = _esc(chunk.source_file)
            location_label = _esc(chunk.location)

            # Try to extract a "name" from the preview for bolding
            name_part, rest_part = self._split_name_from_preview(preview)
            display_html = f"<b>{_esc(name_part)}</b>" if name_part else ""
            if rest_part:
                rest_escaped = _phone_span(_esc(rest_part))
                if display_html:
                    display_html += f"<br>{rest_escaped}"
                else:
                    display_html = rest_escaped

            display_html += (
                f'<br><span style="color:#555;font-size:0.88em;">'
                f'📁 {source_label} | {location_label}</span>'
            )

            html_parts.append(
                '<div class="chatbot-result-card" '
                'style="border-left:3px solid #42a5f5;padding:8px 12px;'
                'margin-bottom:8px;background:#f0f7ff;border-radius:6px;">'
                + display_html
                + '</div>'
            )

        html_parts.append(
            '<br><small style="color:#999;">'
            '💡 <i>Tip: Add area, specialty, or doctor name for more targeted results.</i>'
            '</small>'
        )

        return "".join(html_parts)

    def _split_name_from_preview(self, preview: str) -> Tuple[str, str]:
        """Try to split a doctor/hospital name from the rest of the preview text."""
        # If it's pipe-delimited (from xlsx), use first field as name
        if " | " in preview:
            parts = preview.split(" | ", 1)
            return parts[0].strip(), parts[1].strip() if len(parts) > 1 else ""
        # If starts with "Dr." or a number, take first line
        lines = preview.split("\n", 1)
        if len(lines) > 1:
            return lines[0].strip(), lines[1].strip()
        # Otherwise, just return the whole thing as the name
        if len(preview) < 120:
            return preview, ""
        return "", preview

    # --- PDF extraction ---

    def _extract_pdf_chunks(self, pdf_path: Path, start_id: int) -> List[Chunk]:
        out = []
        reader = PdfReader(str(pdf_path))
        current_id = start_id

        for i, page in enumerate(reader.pages, start=1):
            text = page.extract_text() or ""
            text = text.strip()
            if not text:
                continue

            lines = [re.sub(r"\s+", " ", ln).strip() for ln in text.splitlines() if ln.strip()]
            record_lines = self._group_pdf_lines_into_records(lines)

            if not record_lines:
                record_lines = [re.sub(r"\s+", " ", text)]

            for para_idx, para in enumerate(record_lines, start=1):
                out.append(
                    Chunk(
                        chunk_id=current_id,
                        source_file=pdf_path.name,
                        source_type="pdf",
                        location=f"page {i}, section {para_idx}",
                        text=para,
                    )
                )
                current_id += 1
        return out

    def _group_pdf_lines_into_records(self, lines: List[str]) -> List[str]:
        records: List[str] = []
        current = ""

        for raw_line in lines:
            for line in self._split_inline_records(raw_line):
                if len(line) < 4:
                    continue

                is_new_record = bool(
                    re.match(r"^\d+\s", line)
                    or re.match(r"^[a-z]{2}/\d+", line, flags=re.IGNORECASE)
                    or re.match(r"^(dr\.?|doctor)\b", line, flags=re.IGNORECASE)
                )

                if is_new_record:
                    if current:
                        records.append(current.strip())
                    current = line
                else:
                    if current:
                        current = f"{current} {line}"
                    else:
                        current = line

        if current:
            records.append(current.strip())

        filtered = [r for r in records if len(r) >= 20]
        return filtered

    def _split_inline_records(self, line: str) -> List[str]:
        text = re.sub(r"\s+", " ", line).strip()
        if not text:
            return []

        chunks = re.split(
            r"(?=\s\d{1,3}\.\s)|(?=\s[a-z]{2}/\d{1,4})",
            f" {text}",
            flags=re.IGNORECASE,
        )
        cleaned = [c.strip() for c in chunks if c and c.strip()]
        if not cleaned:
            return [text]
        return cleaned

    # --- XLSX extraction ---

    def _extract_xlsx_chunks(self, xlsx_path: Path, start_id: int) -> List[Chunk]:
        out = []
        workbook = load_workbook(filename=str(xlsx_path), read_only=True, data_only=True)
        current_id = start_id

        for sheet in workbook.worksheets:
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                continue

            headers = [str(h).strip() if h is not None else "" for h in rows[0]]
            for row_idx, row in enumerate(rows[1:], start=2):
                fields = []
                for col_idx, value in enumerate(row):
                    if value is None:
                        continue
                    header = headers[col_idx] if col_idx < len(headers) else f"col_{col_idx + 1}"
                    header = header or f"col_{col_idx + 1}"
                    fields.append(f"{header}: {value}")

                if not fields:
                    continue

                out.append(
                    Chunk(
                        chunk_id=current_id,
                        source_file=xlsx_path.name,
                        source_type="xlsx",
                        location=f"sheet '{sheet.title}', row {row_idx}",
                        text=" | ".join(fields),
                    )
                )
                current_id += 1

        workbook.close()
        return out

    # --- BM25 ---

    def _prepare_bm25_structures(self) -> None:
        self.tokenized_chunks = [tokenize(chunk.text) for chunk in self.chunks]
        self.chunk_term_freqs = [dict(Counter(tokens)) for tokens in self.tokenized_chunks]
        self._all_known_terms = {term for tokens in self.tokenized_chunks for term in tokens}

        doc_freq_counter: Dict[str, int] = defaultdict(int)
        for tokens in self.tokenized_chunks:
            for term in set(tokens):
                doc_freq_counter[term] += 1
        self.doc_freq = dict(doc_freq_counter)

        lengths = [len(tokens) for tokens in self.tokenized_chunks]
        self.avg_doc_len = (sum(lengths) / len(lengths)) if lengths else 0.0

    def _bm25_score(self, query_terms: List[str], chunk_idx: int) -> float:
        k1 = 1.5
        b = 0.75
        tf = self.chunk_term_freqs[chunk_idx]
        doc_len = len(self.tokenized_chunks[chunk_idx]) or 1
        n_docs = len(self.chunks)
        score = 0.0

        for term in query_terms:
            if term not in tf:
                continue
            df = self.doc_freq.get(term, 0)
            if df == 0:
                continue

            idf = math.log(1 + (n_docs - df + 0.5) / (df + 0.5))
            term_freq = tf[term]
            numerator = term_freq * (k1 + 1)
            denominator = term_freq + k1 * (1 - b + b * (doc_len / (self.avg_doc_len or 1)))
            score += idf * (numerator / denominator)
        return score

    def _retrieve(
        self,
        query_terms: List[str],
        required_terms: Set[str],
        phrase_filters: List[str],
        top_k: int = 5,
    ) -> List[Tuple[float, float, Chunk]]:
        scored = []
        min_score = 0.35 if len(required_terms) <= 2 else 1.2
        min_coverage = 1.0 if (len(required_terms) <= 2 and phrase_filters) else 0.5

        for idx, chunk in enumerate(self.chunks):
            score = self._bm25_score(query_terms, idx)
            if score <= 0:
                continue

            chunk_terms = set(self.tokenized_chunks[idx])
            matched_required = sum(1 for t in required_terms if t in chunk_terms)
            coverage = matched_required / max(1, len(required_terms))

            if matched_required == 0:
                continue
            if len(required_terms) >= 2 and coverage < min_coverage:
                continue
            if score < min_score:
                continue
            if phrase_filters:
                lowered_text = normalize_text(chunk.text)
                if not any(phrase in lowered_text for phrase in phrase_filters):
                    continue

            scored.append((score, coverage, chunk))

        scored.sort(key=lambda x: (x[1], x[0]), reverse=True)
        return scored[:top_k]

    # --- Spelling correction ---

    def _correct_terms(self, query_terms: List[str]) -> Tuple[List[str], List[Tuple[str, str]]]:
        corrected = []
        pairs = []
        for term in query_terms:
            if len(term) <= 3:
                corrected.append(term)
                continue
            if term in self._all_known_terms:
                corrected.append(term)
                continue
            candidate = self._closest_term(term)
            fixed = candidate if candidate else term
            corrected.append(fixed)
            if fixed != term:
                pairs.append((term, fixed))
        return corrected, pairs

    def _required_terms(self, query_terms: List[str]) -> Set[str]:
        required = {t for t in query_terms if len(t) >= 3 and not t.isdigit()}
        if not required:
            required = set(query_terms)
        return required

    def _wants_address(self, query_terms: List[str]) -> bool:
        return any(t in {"address", "location", "where"} for t in query_terms)

    def _extract_address_like_text(self, text: str) -> str:
        clean = re.sub(r"\s+", " ", text).strip()

        hospital_match = re.search(
            r"hospital\s+(.*?)\s+(all purpose|all pupose|yes|no|\d{4}\s*-\s*\d{2}\s*-\s*\d{2})",
            clean,
            flags=re.IGNORECASE,
        )
        if hospital_match:
            value = hospital_match.group(1).strip(" ,.-")
            if value:
                return value

        phone_idx = re.search(r"\b(?:\d{3,5}\s*-\s*\d{5,8}|\d{10})\b", clean)
        if phone_idx:
            before_phone = clean[: phone_idx.start()].strip()
            before_phone = re.sub(r"^dr\.?\s+[a-z\s\.]+?\s+", "", before_phone, flags=re.IGNORECASE)
            if before_phone:
                return before_phone.strip(" ,.-")

        return ""

    def _closest_term(self, term: str) -> str:
        if not self._all_known_terms:
            return ""
        best = ""
        best_score = float("inf")
        for known in self._all_known_terms:
            if abs(len(known) - len(term)) > 3:
                continue
            if term and known and term[0] != known[0]:
                continue
            d = self._levenshtein_distance(term, known)
            if d < best_score:
                best_score = d
                best = known
        max_edit = 1 if len(term) <= 5 else 2
        return best if best_score <= max_edit else ""

    def _phrase_filters(self, query: str) -> List[str]:
        q = normalize_text(query)
        if re.search(r"\b(near|nearby|around|in|at|address|location|where)\b", q):
            return []
        phrases = []
        if re.search(r"\b(dr|doctor)\b", q):
            cleaned = re.sub(r"[^a-z0-9\s]", " ", q)
            cleaned = re.sub(r"\s+", " ", cleaned).strip()
            if cleaned:
                words = [
                    w
                    for w in cleaned.split()
                    if w not in INTENT_WORDS and w not in {"of", "for", "what", "is", "the"}
                ]
                cleaned = " ".join(words).strip()
            if cleaned:
                without_titles = re.sub(r"\b(dr|doctor)\b", "", cleaned)
                without_titles = re.sub(r"\s+", " ", without_titles).strip()
                tokens = without_titles.split()
                has_org_hint = any(t in ORG_HINT_WORDS for t in tokens)
                if without_titles and len(tokens) >= 2 and not has_org_hint:
                    phrases.append(cleaned)
                    phrases.append(without_titles)
        return phrases

    @staticmethod
    def _levenshtein_distance(a: str, b: str) -> int:
        if a == b:
            return 0
        if not a:
            return len(b)
        if not b:
            return len(a)

        prev = list(range(len(b) + 1))
        for i, ca in enumerate(a, start=1):
            curr = [i]
            for j, cb in enumerate(b, start=1):
                ins = curr[j - 1] + 1
                dele = prev[j] + 1
                sub = prev[j - 1] + (0 if ca == cb else 1)
                curr.append(min(ins, dele, sub))
            prev = curr
        return prev[-1]

    # --- Index persistence ---

    def _save_index(self) -> None:
        payload = {
            "schema_version": INDEX_SCHEMA_VERSION,
            "chunks": [asdict(chunk) for chunk in self.chunks],
            "doc_freq": self.doc_freq,
            "avg_doc_len": self.avg_doc_len,
        }
        self.index_file.write_text(
            json.dumps(payload, ensure_ascii=True, indent=2), encoding="utf-8"
        )

    def _load_index(self) -> None:
        payload = json.loads(self.index_file.read_text(encoding="utf-8"))
        if payload.get("schema_version") != INDEX_SCHEMA_VERSION:
            raise ValueError("Index schema version mismatch. Rebuild required.")
        self.chunks = [Chunk(**item) for item in payload.get("chunks", [])]
        self.doc_freq = payload.get("doc_freq", {})
        self.avg_doc_len = payload.get("avg_doc_len", 0.0)

        if not self.chunks:
            raise ValueError("Index file exists but has no chunks. Rebuild index.")

        self._prepare_bm25_structures()


# ---------------------------------------------------------------------------
# Global singleton + app.py compatibility layer
# ---------------------------------------------------------------------------

_BOT: Optional[LocalRAGChatbot] = None
_INITIALIZED = False


def init() -> None:
    """Load all Doctor's Data files and build the BM25 index. Called once at startup by app.py."""
    global _BOT, _INITIALIZED

    if _INITIALIZED:
        return

    _BOT = LocalRAGChatbot(data_dir=DATA_DIR, index_file=INDEX_FILE)

    if not DATA_DIR.exists():
        print("[!] Doctor's Data folder not found — chatbot will return no results.")
        _INITIALIZED = True
        return

    try:
        _BOT.load_or_build_index()
        print(f"[chatbot] Loaded {len(_BOT.chunks)} chunks from Doctor's Data. Ready ✓")
    except Exception as e:
        print(f"[chatbot] Error building index: {e}")

    _INITIALIZED = True


def get_response(message: str, context: str = "") -> str:
    """
    Main chatbot function called by app.py.
    Returns an HTML-formatted string ready for the frontend to render.
    If 'context' (report summary) is provided, uses Groq LLM to answer the question based on it.
    """
    if not message or not message.strip():
        return "Please type a message! 😊"

    # 1. If we have report context, try to answer using the LLM first
    if context and context.strip():
        llm_reply = _ask_llm_with_context(message.strip(), context.strip())
        if llm_reply:
            return llm_reply

    # 2. Otherwise fall back to local RAG (Doctor search)

    if not _INITIALIZED:
        init()

    if _BOT is None or not _BOT.chunks:
        # Still try small-talk even if index is empty
        smalltalk = get_smalltalk_reply(message)
        if smalltalk:
            return smalltalk
        return (
            "⚠️ No data is loaded yet. Please ensure the <b>Doctor's Data</b> "
            "folder contains PDF or XLSX files and restart the server."
        )

    try:
        return _BOT.ask(message.strip(), top_k=6)
    except Exception as e:
        return f"⚠️ An error occurred: {_esc(str(e))}"


# ---------------------------------------------------------------------------
# LLM Integration for Report Context
# ---------------------------------------------------------------------------

def _load_api_key() -> str | None:
    """Load the Groq API key from .env file."""
    try:
        from dotenv import load_dotenv
        load_dotenv()
    except ImportError:
        pass
    
    key = os.environ.get("GROQ_API_KEY", "").strip()
    return key if key else None

def _ask_llm_with_context(question: str, context: str) -> str | None:
    """
    Use Groq LLM to answer a user's question based on their medical report summary.
    Returns None if the question doesn't seem related to the report, or if API fails.
    """
    api_key = _load_api_key()
    if not api_key:
        return None  # Fallback to local RAG if no internet/API key
    
    # Simple check: does the question look like it's about the report?
    # If it has intent words for doctors/addresses, likely a local RAG query
    q_norm = normalize_text(question)
    q_words = set(q_norm.split())
    if q_words & INTENT_WORDS and not any(w in q_norm for w in ["my", "report", "values", "score", "risk"]):
        return None # Looks like a doctor search

    try:
        from groq import Groq
        client = Groq(api_key=api_key)
        
        system_prompt = (
            "You are SwasthyaBot, a helpful AI medical assistant. "
            "You have been provided with the user's latest medical report summary. "
            "Answer their question based ONLY on this summary. "
            "Keep the answer short, concise (2-3 sentences max), and easy to understand. "
            "Use HTML formatting (<b>bold</b>, <br> for newlines) instead of markdown. "
            "If the question is completely unrelated to the report or health, say 'I can only answer questions about your report or help you find doctors.' "
            "Be reassuring but remind them to consult a doctor for actual medical advice."
        )
        
        # Trim context to avoid sending massive amounts of text
        trimmed_context = context[:6000]

        chat_completion = client.chat.completions.create(
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": f"Context (My Report Summary):\n{trimmed_context}\n\nMy Question: {question}"},
            ],
            model="llama-3.3-70b-versatile",
            temperature=0.3,
            max_tokens=600,
            top_p=0.9,
        )

        response = chat_completion.choices[0].message.content
        if response:
            return f'<div style="background:#f8f9fa; padding:10px; border-radius:6px; border-left:3px solid #8b5cf6;"><b>Report Q&A:</b><br>{response}</div>'
        return None
    except Exception as e:
        print(f"LLM Context Chat Error: {e}")
        return None

# ---------------------------------------------------------------------------
# CLI mode
# ---------------------------------------------------------------------------

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Local no-paid-API RAG chatbot for doctor PDFs/XLSX in Data folder."
    )
    parser.add_argument(
        "--rebuild",
        action="store_true",
        help="Force rebuild of index_store.json from files in Data folder.",
    )
    return parser.parse_args()


def run_chat(rebuild: bool = False) -> None:
    bot = LocalRAGChatbot(data_dir=DATA_DIR, index_file=INDEX_FILE)
    had_index = INDEX_FILE.exists()
    needs_refresh = bot._needs_rebuild() if had_index else True
    bot.load_or_build_index(force_rebuild=rebuild)

    print("\nDoctor Data Chatbot (local RAG)")
    if rebuild:
        print("Index status: rebuilt from Data folder.")
    elif not had_index:
        print("Index status: created from Data folder.")
    elif needs_refresh:
        print("Index status: refreshed from Data folder updates.")
    else:
        print("Index status: loaded cached index.")
    print("Type your question in natural language.")
    print("Type 'exit' to quit.\n")

    while True:
        try:
            user_input = input("You: ").strip()
        except KeyboardInterrupt:
            print("\nBye.")
            break
        except EOFError:
            print("\nBye.")
            break
        if not user_input:
            continue
        if user_input.lower() in {"exit", "quit"}:
            print("Bye.")
            break

        response = bot.ask(user_input, top_k=5)
        print(f"\nBot:\n{response}\n")


if __name__ == "__main__":
    args = parse_args()
    run_chat(rebuild=args.rebuild)